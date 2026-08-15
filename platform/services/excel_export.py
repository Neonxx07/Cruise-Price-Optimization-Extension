"""Excel export service.

Writes a color-coded, sorted .xlsx report from booking results — the
xlsx counterpart to csv_export.py, for when a spreadsheet is easier to
work with than a raw CSV.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.calculator import total_optimization_savings
from core.models import BookingResult

COLS = [
    "Booking ID", "Cruise Line", "Status", "Confidence",
    "Old Total ($)", "New Total ($)", "Price Drop ($)", "OBC Change ($)",
    "Net Saving ($)", "Category", "New Category", "Note",
    "Lost Packages", "Lost Fares", "Re-addable Fares", "Gained Fares",
    "Checked At",
]

_FILLS = {
    "OPTIMIZATION": PatternFill("solid", fgColor="C6EFCE"),      # green
    # Deliberately distinct from OPTIMIZATION's green: a category upgrade
    # always needs human review before switching (different physical
    # room/deck), unlike a confirmed same-category OPTIMIZATION.
    "UPGRADE_AVAILABLE": PatternFill("solid", fgColor="D9D2E9"), # light purple
    "TRAP": PatternFill("solid", fgColor="FFEB9C"),           # amber
    "NO_SAVING": PatternFill("solid", fgColor="F2F2F2"),      # grey
    "ERROR": PatternFill("solid", fgColor="FFC7CE"),          # red
    "WLT": PatternFill("solid", fgColor="DDEBF7"),            # light blue
    "PAID_IN_FULL": PatternFill("solid", fgColor="DDEBF7"),   # light blue
    "SKIPPED_TODAY": PatternFill("solid", fgColor="DDEBF7"),  # light blue
}
_SORT_ORDER = {
    "OPTIMIZATION": 0, "UPGRADE_AVAILABLE": 1, "TRAP": 2, "WLT": 3,
    "PAID_IN_FULL": 4, "NO_SAVING": 5, "SKIPPED_TODAY": 6, "ERROR": 7,
}
_HDR_FILL = PatternFill("solid", fgColor="1F3864")
_HDR_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_DATA_FONT = Font(name="Calibri", size=10)
_BOLD_FONT = Font(name="Calibri", size=10, bold=True)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_COL_WIDTHS = [14, 12, 14, 10, 13, 13, 13, 13, 13, 10, 12, 30, 24, 24, 24, 24, 20]


def _row(r: BookingResult) -> list:
    return [
        r.booking_id,
        r.cruise_line.value,
        r.status.value,
        r.confidence,
        r.old_total,
        r.new_total,
        r.price_drop,
        r.obc_change,
        r.net_saving,
        r.price_category or "",
        r.new_price_category or "",
        r.note,
        " | ".join(r.lost_pkg_names),
        " | ".join(r.lost_fares),
        " | ".join(r.re_addable_fares),
        " | ".join(r.gained_fares),
        r.checked_at.isoformat() if r.checked_at else "",
    ]


def export_results_excel(results: list[BookingResult], path: str | Path) -> None:
    """Write a color-coded, sorted .xlsx report with a Results + Summary sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    ws.append(COLS)
    for cell in ws[1]:
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = _CENTER
        cell.border = _THIN
    ws.row_dimensions[1].height = 26

    def sort_key(r: BookingResult):
        return (_SORT_ORDER.get(r.status.value, 9), -r.net_saving)

    for r in sorted(results, key=sort_key):
        ws.append(_row(r))
        ri = ws.max_row
        fill = _FILLS.get(r.status.value, _FILLS["NO_SAVING"])
        for ci, cell in enumerate(ws[ri], 1):
            cell.fill = fill
            cell.border = _THIN
            cell.font = _BOLD_FONT if ci in (1, 3, 9) else _DATA_FONT
            cell.alignment = _CENTER if ci in (3, 4, 10, 11) else _LEFT

    for col_letter, width in zip((chr(65 + i) for i in range(len(_COL_WIDTHS))), _COL_WIDTHS):
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Summary sheet
    ss = wb.create_sheet("Summary")
    ss.column_dimensions["A"].width = 24
    ss.column_dimensions["B"].width = 16

    def count(status: str) -> int:
        return sum(1 for r in results if r.status.value == status)

    opts = [r for r in results if r.status.value == "OPTIMIZATION"]
    total_saved = total_optimization_savings(results)

    summary_rows = [
        ("Total Checked", len(results)),
        ("Optimizations", len(opts)),
        ("Upgrades Available", count("UPGRADE_AVAILABLE")),
        ("Traps", count("TRAP")),
        ("WLT", count("WLT")),
        ("Paid In Full", count("PAID_IN_FULL")),
        ("No Saving", count("NO_SAVING")),
        ("Skipped Today", count("SKIPPED_TODAY")),
        ("Errors", count("ERROR")),
        ("Total Savings Found ($)", round(total_saved, 2)),
    ]
    hdr_font = Font(bold=True, name="Calibri")
    for label, value in summary_rows:
        ss.append([label, value])
        ss[f"A{ss.max_row}"].font = hdr_font

    wb.save(path)
